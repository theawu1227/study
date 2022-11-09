# -*-coding:utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
# import redis
import os
import socket
import easygui


class Table():
    def __init__(self, dic):
        self.school = dic['school'].strip()
        self.meeting = dic['meeting'].strip()
        self.theme = dic['theme'].strip()
        self.date_time = dic['date_time'].strip()
        self.speaker = dic['speaker'].strip()
        self.chairman = dic['chairman'].strip()
        self.people_name = dic['people'].strip()
        self.url = dic['url'].strip()
        self.table_name = dic['table_name'].strip()
        name = ("".join([self.people_name,self.chairman,self.speaker])).strip()
        print(name)
        excel_path = 'D:\桌面\我的会议'
        self.excel_file = excel_path + "/" + self.table_name
        isExists = os.path.exists(excel_path)
        if not isExists:
            os.makedirs(excel_path)
            chos = ['新建文件', '已复制指定文件']
            file_l = easygui.buttonbox('已创建文件夹(与run.exe同一文件夹)\n是否新建文件？（或将指定文件放入文件夹table_dir中）', title='创建文件夹', choices=chos)
            if file_l == "已复制指定文件":
                self.input_table()
                return
            elif file_l == "新建文件":
                self.create_table()
                self.input_table()
                return

    def create_table(self):
        meeting_list, date_time_list, chairman_list, speaker_list, school_list, people_name_list, theme_list, list_url = [
            [] for i in range(8)]
        dic = {
            '会议名称': meeting_list,  # 1
            '会议主题': theme_list,  # 2
            '会议时间': date_time_list,  # 3
            '主题演讲人': speaker_list,  # 4
            '会议主席': chairman_list,  # 5
            '参会人': people_name_list,  # 6
            '机构': school_list,  # 7
            '人员列表': list_url
        }
        df = pd.DataFrame(dic)
        df.dropna(axis=0, subset=['会议主席', '主题演讲人', '参会人'], inplace=True)
        df.to_excel(self.excel_file, index=False)
        # print('创建成功')

    def input_table(self):
        # school = self.school.replace('\n', "")
        # from openpyxl import load_workbook
        dic2 = {
            1: self.meeting,  # 会议名称
            2: self.theme,  # 会议主题
            3: self.date_time,  # 会议时间
            4: self.speaker,  # 演讲人
            5: self.chairman,  # 会议主席
            6: self.people_name,  # 参会人
            7: self.school,  # 机构
            8: self.url  # 人员列表链接
        }
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append(dic2)
            wb.save(self.excel_file)
        except Exception as e:
            print(e)
            print('=======')
